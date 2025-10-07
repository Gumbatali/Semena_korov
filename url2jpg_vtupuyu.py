#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скачивает все изображения из Excel (колонка J).
Сохраняет в порядке строк: 001.jpg, 002.jpg, ...
Поддерживает обычные ссылки и гиперссылки Excel.
"""

import os
import requests
from openpyxl import load_workbook
from tqdm import tqdm
from pathlib import Path

# === НАСТРОЙКИ ===
EXCEL_PATH = "links.xlsx"          # путь к Excel-файлу
SHEET_NAME = 0                     # имя или индекс листа (0 = первый)
COLUMN_INDEX = 9                   # J = 10-я колонка, индекс с 0 => 9
OUTPUT_DIR = "downloads"           # куда сохранять
TIMEOUT = 30                       # секунд
# =================

Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
session = requests.Session()
session.headers.update({"User-Agent": "excel-image-downloader"})

# читаем Excel (чтобы работали hyperlink'и)
wb = load_workbook(EXCEL_PATH, read_only=False, data_only=True)
ws = wb.worksheets[SHEET_NAME] if isinstance(SHEET_NAME, int) else wb[SHEET_NAME]

# собираем ссылки
links = []
for row in ws.iter_rows():
    cell = row[COLUMN_INDEX]
    if cell is None:
        continue

    # если гиперссылка Excel
    if cell.hyperlink is not None:
        link = cell.hyperlink.target
    else:
        link = str(cell.value).strip() if cell.value else ""

    if link.lower().startswith("http"):
        links.append(link)

print(f"Найдено {len(links)} ссылок для скачивания.")

# скачивание
for idx, link in enumerate(tqdm(links, desc="Downloading", unit="file")):
    try:
        r = session.get(link, stream=True, timeout=TIMEOUT)
        if r.status_code == 200:
            ctype = r.headers.get("Content-Type", "").lower()
            if "jpeg" in ctype:
                ext = ".jpg"
            elif "png" in ctype:
                ext = ".png"
            elif "webp" in ctype:
                ext = ".webp"
            elif "gif" in ctype:
                ext = ".gif"
            else:
                ext = ".bin"

            filename = f"{idx+1:03d}{ext}"
            path = os.path.join(OUTPUT_DIR, filename)
            with open(path, "wb") as f:
                for chunk in r.iter_content(1024 * 64):
                    f.write(chunk)
        else:
            print(f"\n[!] Ошибка {r.status_code}: {link}")
    except Exception as e:
        print(f"\n[!] Ошибка при скачивании {link}: {e}")

print(f"\n✅ Готово. Файлы сохранены в папке: {os.path.abspath(OUTPUT_DIR)}")
