#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скачивает фотографии из Excel в две папки kartinki/1 и kartinki/2.
0 не скачивает.
Названия файлов: ObservationId_<номер_строки>.jpg
"""

import os
import re
import requests
from pathlib import Path
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

# === НАСТРОЙКИ ===
EXCEL_PATH = "links.xlsx"
OUTPUT_DIR = "kartinki"
SHEET_NAME = 0
COLUMN_ID = 0          # ObservationId (A)
COLUMN_USAGE = 8       # IsUsedForAnalysis (I)
COLUMN_LINK = 9        # PhotoLink (J)
MAX_WORKERS = 20       # параллельных загрузок
TIMEOUT = 20           # секунд
# =================

# Настраиваем окружение
session = requests.Session()
session.headers.update({"User-Agent": "fast-image-downloader"})
Path(OUTPUT_DIR, "1").mkdir(parents=True, exist_ok=True)
Path(OUTPUT_DIR, "2").mkdir(parents=True, exist_ok=True)

# Загружаем Excel
wb = load_workbook(EXCEL_PATH, read_only=False, data_only=True)
ws = wb.worksheets[SHEET_NAME] if isinstance(SHEET_NAME, int) else wb[SHEET_NAME]

# Собираем список ссылок
records = []
for i, row in enumerate(ws.iter_rows()):
    if i == 0:  # пропуск заголовка
        continue
    if len(row) <= COLUMN_LINK:
        continue

    usage = str(row[COLUMN_USAGE].value).strip() if row[COLUMN_USAGE].value else "2"
    if usage == "0":
        continue  # пропускаем

    obs = str(row[COLUMN_ID].value).strip() if row[COLUMN_ID].value else f"noid_{i+1}"

    cell = row[COLUMN_LINK]
    link = ""
    if cell and getattr(cell, "hyperlink", None):
        link = cell.hyperlink.target
    elif cell and cell.value:
        link = str(cell.value).strip()

    if not link.lower().startswith("http"):
        continue

    records.append((obs, link, usage, i + 1))

print(f"Найдено {len(records)} ссылок для скачивания (1 и 2).")

# --- функция для скачивания ---
def download_one(rec):
    obs, link, usage, rownum = rec
    folder = "1" if usage == "1" else "2"
    out_dir = Path(OUTPUT_DIR, folder)
    filename = f"{obs}_{rownum}.jpg"
    filepath = out_dir / filename

    # пропускаем, если файл уже есть
    if filepath.exists():
        return {"ok": True, "obs": obs}

    try:
        r = session.get(link, stream=True, timeout=TIMEOUT)
        if r.status_code == 200:
            with open(filepath, "wb") as f:
                for chunk in r.iter_content(1024 * 64):
                    if chunk:
                        f.write(chunk)
            return {"ok": True, "obs": obs}
        else:
            return {"ok": False, "obs": obs, "msg": f"HTTP {r.status_code}"}
    except Exception as e:
        return {"ok": False, "obs": obs, "msg": str(e)}

# --- параллельная загрузка ---
with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
    futures = [ex.submit(download_one, rec) for rec in records]
    for f in tqdm(as_completed(futures), total=len(futures), desc="Downloading", unit="file"):
        res = f.result()
        if not res["ok"]:
            print(f"\n[!] Ошибка при скачивании {res['obs']}: {res.get('msg','')}")

print(f"\n✅ Готово! Все изображения сохранены в '{os.path.abspath(OUTPUT_DIR)}'")
